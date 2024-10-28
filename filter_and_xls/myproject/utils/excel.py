# utils/excel.py

import openpyxl
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import transaction
from django.db.models import ForeignKey, BooleanField
from django.http import HttpResponse
from myapp.schemas import ApiResponse 



def import_from_excel(model, excel_file, exclude_fields=None):
    if exclude_fields is None:
        exclude_fields = []

    meta = model._meta
    model_field_names = [field.name for field in meta.fields if field.name not in exclude_fields]

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]  # Перша строка - заголовки
    errors = []
    objects_to_create = []
    objects_to_update = []

    # Створюємо мапінг заголовків до полів моделі
    header_field_map = {}
    for header in headers:
        if header in model_field_names:
            header_field_map[header] = header
        else:
            # Спробуємо знайти поле моделі з таким verbose_name
            for field in meta.fields:
                if field.verbose_name == header and field.name not in exclude_fields:
                    header_field_map[header] = field.name
                    break
            else:
                errors.append(f"Заголовок '{header}' не відповідає жодному полю моделі.")

    if errors:
        return ApiResponse(data={}, errors=errors)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        data = {}
        for idx, cell in enumerate(row):
            header = headers[idx]
            if header not in header_field_map:
                continue  # Ігноруємо поля, які не є полями моделі

            field_name = header_field_map[header]
            field = meta.get_field(field_name)
            value = cell.value

            # Обробка BooleanField
            if isinstance(field, BooleanField):
                if value == "Так":
                    value = True
                elif value == "Ні":
                    value = False
                elif value is None or value == "":
                    errors.append(f"Рядок {row_idx}: Поле '{field_name}' не може бути пустим.")
                    continue
                else:
                    errors.append(f"Рядок {row_idx}: Некоректне значення '{value}' для поля '{field_name}'. Очікується 'Так' або 'Ні'.")
                    continue

            # Обробка ForeignKey
            if isinstance(field, ForeignKey):
                rel_model = field.related_model
                try:
                    rel_obj = rel_model.objects.get(pk=value)
                    data[field_name] = rel_obj
                except ObjectDoesNotExist:
                    errors.append(f"Рядок {row_idx}: Не знайдено {field_name} з ID {value}")
            else:
                data[field_name] = value

        # Обробка поля 'id' для оновлення існуючих об'єктів
        id_value = data.get('id')
        try:
            if id_value and id_value != 0:
                # Оновлюємо існуючий об'єкт
                try:
                    obj = model.all_objects.get(pk=id_value)
                    for attr, value in data.items():
                        if attr != 'id':  # Не змінюємо поле 'id'
                            setattr(obj, attr, value)
                    obj.full_clean()
                    objects_to_update.append(obj)
                except ObjectDoesNotExist:
                    errors.append(f"Рядок {row_idx}: Об'єкт з ID {id_value} не знайдено для оновлення")
            else:
                # Створюємо новий об'єкт
                data.pop('id', None)  # Видаляємо 'id', якщо він є
                obj = model(**data)
                obj.full_clean()
                objects_to_create.append(obj)
        except ValidationError as ve:
            errors.append(f"Рядок {row_idx}: {ve.messages}")

    if errors:
        return ApiResponse(data={}, errors=errors)

    # Зберігаємо дані в базу в межах транзакції
    with transaction.atomic():
        if objects_to_create:
            model.objects.bulk_create(objects_to_create)
        if objects_to_update:
            for obj in objects_to_update:
                obj.save()

    return ApiResponse(
        data={"created": len(objects_to_create), "updated": len(objects_to_update)},
        meta={"message": "Дані успішно імпортовано!"},
        errors=[]
    )





def export_to_excel(queryset, exclude_fields=None):
    if exclude_fields is None:
        exclude_fields = []

    model = queryset.model
    meta = model._meta
    fields = [field for field in meta.fields if field.name not in exclude_fields]
    field_names = [field.name for field in fields]
    headers = [field.verbose_name for field in fields]

    wb = openpyxl.Workbook()
    ws = wb.active

    # Додаємо заголовки
    ws.append(headers)

    # Додаємо дані
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field.name)
            # Якщо поле ForeignKey, отримуємо значення його primary key
            if isinstance(field, ForeignKey):
                value = value.pk if value else None
            elif isinstance(field, BooleanField):
                # Конвертуємо булеве значення у "Так"/"Ні"
                value = "Так" if value else "Ні"
            row.append(value)
        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{meta.verbose_name_plural}.xlsx"'
    wb.save(response)
    return response